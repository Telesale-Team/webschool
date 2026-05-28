import sys
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Avg, Max, Min
from django.contrib import messages as django_messages
from django.http import HttpResponse
from .models import Question, Category, ExamConfig, Choice, QuestionParameter
from accounts.models import Student, Teacher
from exams.models import ExamSession, StudentAnswer


def _get_thai_fonts():
    if sys.platform == 'win32':
        return 'C:/Windows/Fonts/tahoma.ttf', 'C:/Windows/Fonts/tahomabd.ttf'
    return (
        '/usr/share/fonts/truetype/tlwg/Garuda.ttf',
        '/usr/share/fonts/truetype/tlwg/Garuda-Bold.ttf',
    )


def home(request):
    from .models import ExamConfig
    config = ExamConfig.get()
    return render(request, 'home.html', {'config': config})


def dashboard(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])
    total_questions = Question.objects.filter(is_active=True).count()
    total_students = Student.objects.count()
    total_sessions = ExamSession.objects.count()
    recent_questions = Question.objects.select_related('category').order_by('-created_at')[:10]
    recent_sessions = ExamSession.objects.select_related('student').order_by('-started_at')[:10]
    return render(request, 'dashboard/index.html', {
        'teacher': teacher,
        'total_questions': total_questions,
        'total_students': total_students,
        'total_sessions': total_sessions,
        'recent_questions': recent_questions,
        'recent_sessions': recent_sessions,
    })


def student_list(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])
    students = Student.objects.annotate(exam_sessions_count=Count('examsession')).order_by('-created_at')
    return render(request, 'dashboard/students.html', {'students': students, 'teacher': teacher})


def question_list(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])
    questions = Question.objects.select_related('parameter', 'category').order_by('number')

    # Stats
    total = questions.count()
    active = questions.filter(is_active=True).count()
    has_answer = QuestionParameter.objects.exclude(correct_answer='').count()
    no_answer = total - has_answer

    param_stats = QuestionParameter.objects.filter(difficulty__isnull=False).aggregate(
        avg_b=Avg('difficulty'), max_b=Max('difficulty'), min_b=Min('difficulty')
    )
    easy = QuestionParameter.objects.filter(difficulty__lt=0.8).count()
    medium = QuestionParameter.objects.filter(difficulty__gte=0.8, difficulty__lt=1.5).count()
    hard = QuestionParameter.objects.filter(difficulty__gte=1.5).count()

    return render(request, 'dashboard/questions.html', {
        'questions': questions,
        'teacher': teacher,
        'total': total,
        'active': active,
        'has_answer': has_answer,
        'no_answer': no_answer,
        'avg_b': round(param_stats['avg_b'], 2) if param_stats['avg_b'] else None,
        'max_b': round(param_stats['max_b'], 2) if param_stats['max_b'] else None,
        'min_b': round(param_stats['min_b'], 2) if param_stats['min_b'] else None,
        'easy': easy,
        'medium': medium,
        'hard': hard,
    })


def edit_question(request, pk):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])
    question = get_object_or_404(Question, pk=pk)
    choices = {c.number: c.body for c in question.choices.all()}
    if request.method == 'POST':
        question.stem = request.POST.get('stem', '').strip()
        question.is_active = request.POST.get('is_active') == 'on'
        if 'image' in request.FILES:
            question.image = request.FILES['image']
        question.save()
        param, _ = QuestionParameter.objects.get_or_create(question=question)
        param.correct_answer = request.POST.get('correct_answer', '').strip()
        param.save()
        # อัพเดท choices
        for n in range(1, 6):
            body = request.POST.get(f'choice_{n}', '').strip()
            Choice.objects.update_or_create(
                question=question, number=n,
                defaults={'body': body}
            )
        django_messages.success(request, f'บันทึกข้อที่ {question.number} เรียบร้อย')
        return redirect('questions:question_list')
    return render(request, 'dashboard/edit_question.html', {
        'teacher': teacher,
        'question': question,
        'choices': choices,
    })


def student_detail(request, pk):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])
    student = get_object_or_404(Student, pk=pk)
    sessions = ExamSession.objects.filter(student=student).order_by('-started_at').prefetch_related('answers__question__parameter')

    session_data = []
    all_scores = []
    for s in sessions:
        answers = s.answers.all()  # ใช้ prefetched cache จาก prefetch_related ด้านบน
        total = len(s.question_order) if s.question_order else 0
        correct = sum(1 for a in answers if a.question.correct_answer and a.answer == a.question.correct_answer)
        score_pct = round(correct / total * 100, 1) if total > 0 else 0
        duration_min = round((s.finished_at - s.started_at).total_seconds() / 60, 1) if s.finished_at else None
        all_scores.append(score_pct)

        # สร้าง answer detail แต่ละข้อ
        answer_details = []
        for a in answers:
            answer_details.append({
                'question': a.question,
                'answer': a.answer,
                'correct_answer': a.question.correct_answer,
                'is_correct': a.question.correct_answer and a.answer == a.question.correct_answer,
            })
        # เรียงตามเลขข้อ
        answer_details.sort(key=lambda x: x['question'].number)

        session_data.append({
            'session': s,
            'correct': correct,
            'total': total,
            'score_pct': score_pct,
            'duration_min': duration_min,
            'answer_details': answer_details,
            'is_finished': s.finished_at is not None,
        })

    best_score = max(all_scores) if all_scores else 0
    avg_score = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0

    return render(request, 'dashboard/student_detail.html', {
        'teacher': teacher,
        'student': student,
        'session_data': session_data,
        'best_score': best_score,
        'avg_score': avg_score,
        'total_sessions': len(session_data),
    })


def edit_student(request, pk):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.first_name = request.POST.get('first_name', student.first_name).strip()
        student.last_name = request.POST.get('last_name', student.last_name).strip()
        student.phone = request.POST.get('phone', student.phone).strip()
        student.grade = request.POST.get('grade', student.grade).strip()
        new_password = request.POST.get('new_password', '').strip()
        if new_password:
            student.set_password(new_password)
        student.save()
        django_messages.success(request, 'อัพเดตข้อมูลนักเรียนเรียบร้อย')
        return redirect('questions:student_list')
    return render(request, 'dashboard/edit_student.html', {
        'teacher': teacher,
        'student': student,
    })


def profile_view(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_info':
            teacher.first_name = request.POST.get('first_name', teacher.first_name).strip()
            teacher.last_name = request.POST.get('last_name', teacher.last_name).strip()
            teacher.email = request.POST.get('email', '').strip()
            teacher.phone = request.POST.get('phone', teacher.phone).strip()
            if 'avatar' in request.FILES:
                teacher.avatar = request.FILES['avatar']
            teacher.save()
            django_messages.success(request, 'อัปเดตข้อมูลเรียบร้อยแล้ว')

        elif action == 'change_password':
            old_pw = request.POST.get('old_password', '')
            new_pw = request.POST.get('new_password', '')
            confirm_pw = request.POST.get('confirm_password', '')
            if not teacher.check_password(old_pw):
                django_messages.error(request, 'รหัสผ่านเดิมไม่ถูกต้อง')
            elif new_pw != confirm_pw:
                django_messages.error(request, 'รหัสผ่านใหม่ไม่ตรงกัน')
            elif len(new_pw) < 4:
                django_messages.error(request, 'รหัสผ่านต้องมีอย่างน้อย 4 ตัวอักษร')
            else:
                teacher.set_password(new_pw)
                teacher.save()
                django_messages.success(request, 'เปลี่ยนรหัสผ่านเรียบร้อยแล้ว')

        return redirect('questions:profile')

    return render(request, 'dashboard/profile.html', {'teacher': teacher})


def report_view(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])

    # --- 1. ภาพรวมผลการสอบ ---
    finished_sessions = ExamSession.objects.filter(finished_at__isnull=False)

    session_scores = []
    for s in finished_sessions.prefetch_related('answers__question__parameter'):
        answers = s.answers.select_related('question__parameter').all()
        total = len(s.question_order) if s.question_order else 0
        correct = sum(
            1 for a in answers
            if a.question.correct_answer and a.answer == a.question.correct_answer
        )
        score_pct = round((correct / total * 100), 1) if total > 0 else 0
        duration_min = (
            round((s.finished_at - s.started_at).total_seconds() / 60, 1)
            if s.finished_at else None
        )
        session_scores.append({
            'session': s,
            'student': s.student,
            'correct': correct,
            'total': total,
            'score_pct': score_pct,
            'duration_min': duration_min,
        })

    scores_list = [s['score_pct'] for s in session_scores]
    avg_score = round(sum(scores_list) / len(scores_list), 1) if scores_list else 0
    max_score = max(scores_list) if scores_list else 0
    min_score = min(scores_list) if scores_list else 0

    # histogram แบ่ง 10 ช่วง (0-9, 10-19, ..., 90-100)
    histogram = [0] * 10
    for sc in scores_list:
        idx = min(int(sc // 10), 9)
        histogram[idx] += 1

    # --- 2. Item Analysis — % ตอบถูกแต่ละข้อ ---
    all_answers = StudentAnswer.objects.select_related('question__parameter').all()
    question_stats = {}
    for ans in all_answers:
        qid = ans.question_id
        if qid not in question_stats:
            question_stats[qid] = {
                'question': ans.question,
                'total': 0,
                'correct': 0,
                'choices': {},
            }
        question_stats[qid]['total'] += 1
        if ans.question.correct_answer and ans.answer == ans.question.correct_answer:
            question_stats[qid]['correct'] += 1
        choice_key = ans.answer if ans.answer else 'ไม่ตอบ'
        question_stats[qid]['choices'][choice_key] = (
            question_stats[qid]['choices'].get(choice_key, 0) + 1
        )

    item_analysis = []
    for qid, stat in question_stats.items():
        pct = round(stat['correct'] / stat['total'] * 100, 1) if stat['total'] > 0 else 0
        item_analysis.append({
            'question': stat['question'],
            'total_attempts': stat['total'],
            'correct': stat['correct'],
            'correct_pct': pct,
            'choices': stat['choices'],
        })
    # เรียงจากยากไปง่าย (% ถูกน้อยไปมาก)
    item_analysis.sort(key=lambda x: x['correct_pct'])
    hardest_10 = item_analysis[:10]

    # --- 3. Leaderboard — best score per student ---
    student_best = {}
    for s in session_scores:
        sid = s['student'].id
        if sid not in student_best or s['score_pct'] > student_best[sid]['score_pct']:
            student_best[sid] = s
    leaderboard = sorted(
        student_best.values(),
        key=lambda x: (-x['score_pct'], x['duration_min'] or 9999)
    )[:20]

    total_students = Student.objects.count()
    students_tested = len(set(s['session'].student_id for s in session_scores))

    return render(request, 'dashboard/reports.html', {
        'teacher': teacher,
        'total_sessions': len(session_scores),
        'total_students': total_students,
        'students_tested': students_tested,
        'avg_score': avg_score,
        'max_score': max_score,
        'min_score': min_score,
        'histogram': histogram,
        'hardest_10': hardest_10,
        'leaderboard': leaderboard,
        'item_analysis': item_analysis,
    })


def _get_report_data():
    """helper — คำนวณข้อมูลรายงาน ใช้ร่วมกันระหว่าง report_view / export views"""
    finished_sessions = ExamSession.objects.filter(finished_at__isnull=False)
    session_scores = []
    for s in finished_sessions.prefetch_related('answers__question__parameter'):
        answers = s.answers.select_related('question__parameter').all()
        total = len(s.question_order) if s.question_order else 0
        correct = sum(1 for a in answers if a.question.correct_answer and a.answer == a.question.correct_answer)
        score_pct = round((correct / total * 100), 1) if total > 0 else 0
        duration_min = round((s.finished_at - s.started_at).total_seconds() / 60, 1) if s.finished_at else None
        session_scores.append({'session': s, 'student': s.student, 'correct': correct, 'total': total, 'score_pct': score_pct, 'duration_min': duration_min})

    scores_list = [s['score_pct'] for s in session_scores]
    all_answers = StudentAnswer.objects.select_related('question__parameter').all()
    question_stats = {}
    for ans in all_answers:
        qid = ans.question_id
        if qid not in question_stats:
            question_stats[qid] = {'question': ans.question, 'total': 0, 'correct': 0}
        question_stats[qid]['total'] += 1
        if ans.question.correct_answer and ans.answer == ans.question.correct_answer:
            question_stats[qid]['correct'] += 1

    item_analysis = []
    for qid, stat in question_stats.items():
        pct = round(stat['correct'] / stat['total'] * 100, 1) if stat['total'] > 0 else 0
        item_analysis.append({'question': stat['question'], 'total_attempts': stat['total'], 'correct': stat['correct'], 'correct_pct': pct})
    item_analysis.sort(key=lambda x: x['correct_pct'])

    student_best = {}
    for s in session_scores:
        sid = s['student'].id
        if sid not in student_best or s['score_pct'] > student_best[sid]['score_pct']:
            student_best[sid] = s
    leaderboard = sorted(student_best.values(), key=lambda x: (-x['score_pct'], x['duration_min'] or 9999))[:20]

    return {
        'session_scores': session_scores,
        'scores_list': scores_list,
        'item_analysis': item_analysis,
        'leaderboard': leaderboard,
        'avg_score': round(sum(scores_list) / len(scores_list), 1) if scores_list else 0,
        'max_score': max(scores_list) if scores_list else 0,
        'min_score': min(scores_list) if scores_list else 0,
    }


def export_questions(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    fmt = request.GET.get('fmt', 'excel')
    questions = Question.objects.select_related('category').order_by('number')

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'คลังข้อสอบ'
        headers = ['ข้อที่', 'โจทย์', 'เฉลย', 'ความยาก', 'สถานะ']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1B6FA8')
            cell.alignment = Alignment(horizontal='center')
        for q in questions:
            ws.append([q.number, (q.stem or q.body_clean)[:100], q.correct_answer or '-', q.difficulty if q.difficulty is not None else '-', 'เปิด' if q.is_active else 'ปิด'])
        ws.column_dimensions['B'].width = 80
        for col in ['A', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 14
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="questions.xlsx"'
        wb.save(response)
        return response

    # PDF
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _font, _font_bold = _get_thai_fonts()
    pdfmetrics.registerFont(TTFont('Tahoma', _font))
    pdfmetrics.registerFont(TTFont('Tahoma-Bold', _font_bold))
    thai = ParagraphStyle('Thai', fontName='Tahoma', fontSize=8, leading=12)
    thai_h1 = ParagraphStyle('ThaiH1', fontName='Tahoma-Bold', fontSize=14, leading=20, spaceAfter=10, alignment=1)
    thai_bold = ParagraphStyle('ThaiBold', fontName='Tahoma-Bold', fontSize=11, leading=16, spaceAfter=6)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = [Paragraph('คลังข้อสอบ', thai_h1), Spacer(1, 0.3*cm)]
    rows = [[str(q.number), Paragraph((q.stem or q.body_clean)[:80], thai), q.correct_answer or '-', str(q.difficulty or '-'), 'เปิด' if q.is_active else 'ปิด'] for q in questions]
    t = Table([['ข้อที่', 'โจทย์', 'เฉลย', 'ความยาก', 'สถานะ']] + rows, colWidths=[1.2*cm, 11*cm, 1.5*cm, 2*cm, 1.8*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Tahoma'), ('FONTNAME', (0, 0), (-1, 0), 'Tahoma-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B6FA8')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('FONTSIZE', (0, 0), (-1, -1), 8), ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3), ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="questions.pdf"'
    return response


def export_students(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.db.models import Count
    fmt = request.GET.get('fmt', 'excel')
    students = Student.objects.annotate(exam_sessions_count=Count('examsession')).order_by('first_name')

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'รายชื่อนักเรียน'
        headers = ['ลำดับ', 'เลขประจำตัว', 'ชื่อ', 'นามสกุล', 'ชั้นปี', 'เบอร์โทร', 'จำนวนครั้งที่สอบ', 'วันที่ลงทะเบียน']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F6B38')
            cell.alignment = Alignment(horizontal='center')
        for i, s in enumerate(students, 1):
            ws.append([i, s.student_id, s.first_name, s.last_name, s.grade, s.phone or '-', s.exam_sessions_count, s.created_at.strftime('%d/%m/%Y') if s.created_at else ''])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
        wb.save(response)
        return response

    # PDF
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _font, _font_bold = _get_thai_fonts()
    pdfmetrics.registerFont(TTFont('Tahoma', _font))
    pdfmetrics.registerFont(TTFont('Tahoma-Bold', _font_bold))
    thai_h1 = ParagraphStyle('ThaiH1', fontName='Tahoma-Bold', fontSize=14, leading=20, spaceAfter=10, alignment=1)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = [Paragraph('รายชื่อนักเรียน', thai_h1), Spacer(1, 0.3*cm)]
    rows = [[str(i), s.student_id, s.first_name, s.last_name, s.grade, s.phone or '-', str(s.exam_sessions_count), s.created_at.strftime('%d/%m/%Y') if s.created_at else ''] for i, s in enumerate(students, 1)]
    t = Table([['#', 'รหัส', 'ชื่อ', 'นามสกุล', 'ชั้นปี', 'เบอร์โทร', 'ครั้งที่สอบ', 'วันที่สมัคร']] + rows, colWidths=[0.8*cm, 2*cm, 3*cm, 3*cm, 1.5*cm, 2.5*cm, 2*cm, 2.7*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Tahoma'), ('FONTNAME', (0, 0), (-1, 0), 'Tahoma-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F6B38')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('FONTSIZE', (0, 0), (-1, -1), 9), ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="students.pdf"'
    return response


def export_excel(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    section = request.GET.get('section', 'all')
    data = _get_report_data()
    wb = openpyxl.Workbook()

    def style_header(ws, color):
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=color)
            cell.alignment = Alignment(horizontal='center')

    if section in ('leaderboard', 'all'):
        ws = wb.active if section == 'leaderboard' else wb.create_sheet('Leaderboard')
        ws.title = 'Leaderboard'
        ws.append(['อันดับ', 'ชื่อ-นามสกุล', 'ชั้นปี', 'คะแนนสูงสุด (%)', 'เวลา (นาที)'])
        style_header(ws, '1F6B38')
        for i, s in enumerate(data['leaderboard'], 1):
            ws.append([i, f"{s['student'].first_name} {s['student'].last_name}", getattr(s['student'], 'grade', ''), s['score_pct'], s['duration_min']])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    if section in ('item_analysis', 'hardest10', 'all'):
        rows = data['item_analysis'][:10] if section == 'hardest10' else data['item_analysis']
        title = '10 ข้อยากสุด' if section == 'hardest10' else 'Item Analysis'
        ws = wb.active if section in ('item_analysis', 'hardest10') else wb.create_sheet(title)
        ws.title = title
        ws.append(['ข้อที่', 'โจทย์', 'เฉลย', 'ตอบถูก', 'ทั้งหมด', '% ถูก'])
        style_header(ws, '7B3F00' if section == 'hardest10' else '375623')
        for item in rows:
            q = item['question']
            ws.append([q.number, (q.stem or q.body)[:80], q.correct_answer, item['correct'], item['total_attempts'], item['correct_pct']])
        ws.column_dimensions['B'].width = 60
        for col in ['A', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 14

    if section == 'all':
        # Sheet ผลการสอบรายบุคคล (ใส่เป็น sheet แรก)
        ws_scores = wb.create_sheet('ผลการสอบ', 0)
        ws_scores.append(['ลำดับ', 'ชื่อ-นามสกุล', 'ชั้นปี', 'คะแนน (%)', 'ถูก', 'ทั้งหมด', 'เวลา (นาที)', 'วันที่สอบ'])
        style_header(ws_scores, '1F497D')
        for i, s in enumerate(data['session_scores'], 1):
            ws_scores.append([i, f"{s['student'].first_name} {s['student'].last_name}", getattr(s['student'], 'grade', ''), s['score_pct'], s['correct'], s['total'], s['duration_min'], s['session'].started_at.strftime('%d/%m/%Y %H:%M') if s['session'].started_at else ''])
        for col in ws_scores.columns:
            ws_scores.column_dimensions[col[0].column_letter].width = 18
        # ลบ default sheet ว่าง
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    filename_map = {'leaderboard': 'leaderboard', 'item_analysis': 'item_analysis', 'hardest10': 'hardest10', 'all': 'report_all'}
    filename = f"{filename_map.get(section, 'report')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_pdf(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')

    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    _font, _font_bold = _get_thai_fonts()
    pdfmetrics.registerFont(TTFont('Tahoma', _font))
    pdfmetrics.registerFont(TTFont('Tahoma-Bold', _font_bold))

    section = request.GET.get('section', 'all')
    data = _get_report_data()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    thai = ParagraphStyle('Thai', fontName='Tahoma', fontSize=10, leading=16)
    thai_bold = ParagraphStyle('ThaiBold', fontName='Tahoma-Bold', fontSize=13, leading=20, spaceAfter=6)
    thai_h1 = ParagraphStyle('ThaiH1', fontName='Tahoma-Bold', fontSize=16, leading=24, spaceAfter=12, alignment=1)

    def make_table(header, rows, col_widths, header_color):
        t = Table([header] + rows, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Tahoma'),
            ('FONTNAME', (0, 0), (-1, 0), 'Tahoma-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_color)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        return t

    elements = []
    title_map = {'leaderboard': 'อันดับนักเรียน (Top 20)', 'hardest10': '10 ข้อที่ยากที่สุด', 'item_analysis': 'Item Analysis ทุกข้อ', 'all': 'รายงานผลการสอบออนไลน์'}
    elements.append(Paragraph(title_map.get(section, 'รายงานผลการสอบออนไลน์'), thai_h1))
    elements.append(Spacer(1, 0.4*cm))

    if section in ('leaderboard', 'all'):
        if section == 'all':
            elements.append(Paragraph('อันดับนักเรียน (Top 20)', thai_bold))
        lb_rows = [[str(i), f"{s['student'].first_name} {s['student'].last_name}", getattr(s['student'], 'grade', '-'), f"{s['score_pct']}%", str(s['duration_min'] or '-')] for i, s in enumerate(data['leaderboard'], 1)]
        elements.append(make_table(['อันดับ', 'ชื่อ-นามสกุล', 'ชั้นปี', 'คะแนน (%)', 'เวลา (นาที)'], lb_rows, [1.5*cm, 6*cm, 2.5*cm, 3*cm, 3*cm], '#1F6B38'))
        elements.append(Spacer(1, 0.5*cm))

    if section in ('hardest10', 'item_analysis', 'all'):
        rows_data = data['item_analysis'][:10] if section == 'hardest10' else data['item_analysis']
        if section == 'all':
            elements.append(Paragraph('10 ข้อที่ยากที่สุด' if section == 'hardest10' else 'Item Analysis', thai_bold))
        ia_rows = [[str(item['question'].number), Paragraph((item['question'].stem or item['question'].body)[:60], thai), item['question'].correct_answer or '-', f"{item['correct']}/{item['total_attempts']}", f"{item['correct_pct']}%"] for item in rows_data]
        color = '#7B3F00' if section == 'hardest10' else '#375623'
        elements.append(make_table(['ข้อที่', 'โจทย์', 'เฉลย', 'ถูก/ทั้งหมด', '% ถูก'], ia_rows, [1.5*cm, 8*cm, 1.5*cm, 2.5*cm, 2.5*cm], color))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    return response


def public_results(request):
    config = ExamConfig.get()
    if not config.is_results_published:
        return render(request, 'public_results_closed.html', {'config': config})

    finished_sessions = ExamSession.objects.filter(
        finished_at__isnull=False
    ).select_related('student').prefetch_related('answers__question__parameter')

    # คำนวณคะแนน best score per student
    student_best = {}
    for s in finished_sessions:
        answers = s.answers.select_related('question__parameter').all()
        total = len(s.question_order) if s.question_order else 0
        correct = sum(1 for a in answers if a.question.correct_answer and a.answer == a.question.correct_answer)
        score_pct = round(correct / total * 100, 1) if total > 0 else 0
        duration_min = round((s.finished_at - s.started_at).total_seconds() / 60, 1) if s.finished_at else None

        sid = s.student_id
        if sid not in student_best or score_pct > student_best[sid]['score_pct']:
            student_best[sid] = {
                'student': s.student,
                'score_pct': score_pct,
                'correct': correct,
                'total': total,
                'duration_min': duration_min,
                'exam_date': s.started_at,
            }

    results = sorted(student_best.values(), key=lambda x: (-x['score_pct'], x['duration_min'] or 9999))

    return render(request, 'public_results.html', {
        'config': config,
        'results': results,
        'total_participants': len(results),
    })


def settings_view(request):
    if not request.session.get('teacher_id'):
        return redirect('accounts:teacher_login')
    teacher = Teacher.objects.get(id=request.session['teacher_id'])
    config = ExamConfig.get()
    if request.method == 'POST':
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile

        config.subject_name = request.POST.get('subject_name', config.subject_name).strip()
        config.year = request.POST.get('year', config.year).strip()
        config.num_choices = int(request.POST.get('num_choices', config.num_choices))
        config.time_limit = int(request.POST.get('time_limit', 0) or 0) * 60

        # Background fields — ถ้ามี upload ให้ใช้ไฟล์ ถ้าไม่มีใช้ URL ที่กรอก
        bg_fields = ['home_bg', 'student_login_bg', 'teacher_login_bg',
                     'student_register_bg', 'teacher_register_bg', 'exam_bg']
        for field in bg_fields:
            upload_key = f'{field}_upload'
            if upload_key in request.FILES:
                f = request.FILES[upload_key]
                path = default_storage.save(f'backgrounds/{f.name}', ContentFile(f.read()))
                setattr(config, field, default_storage.url(path))
            else:
                setattr(config, field, request.POST.get(field, '').strip())

        config.is_results_published = request.POST.get('is_results_published') == 'on'
        config.questions_per_session = int(request.POST.get('questions_per_session', 0) or 0)
        config.question_order = request.POST.get('question_order', 'difficulty')
        config.allow_retake = request.POST.get('allow_retake') == 'on'
        config.save()
        django_messages.success(request, 'บันทึกการตั้งค่าเรียบร้อยแล้ว')
        return redirect('questions:settings')
    return render(request, 'dashboard/settings.html', {
        'teacher': teacher,
        'config': config,
    })
