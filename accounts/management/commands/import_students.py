import re
import openpyxl
from django.core.management.base import BaseCommand
from accounts.models import Student


class Command(BaseCommand):
    help = 'Import ม.6 students from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Path to Excel file')
        parser.add_argument('--sheet', default='ม. 6_169', help='Sheet name (default: ม. 6_169)')
        parser.add_argument('--grade-prefix', default='6', help='Grade prefix to filter (default: 6)')

    def handle(self, *args, **options):
        path = options['excel_path']
        sheet_name = options['sheet']

        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name]

        current_room = ''
        created = 0
        skipped = 0
        seen_ids = set()

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            val0, val1, val2, val3 = row[0], row[1], row[2], row[3]

            # detect classroom header
            if val0 and isinstance(val0, str) and 'ชั้นมัธยมศึกษาปีที่' in val0:
                m = re.search(r'(\d+/\d+)', val0)
                if m:
                    current_room = 'ม.' + m.group(1)
                continue

            # student row: เลขที่=int, เลขประจำตัว=int, ชื่อ=str
            if not (isinstance(val0, int) and isinstance(val1, int) and isinstance(val3, str)):
                continue

            student_id = str(val1)
            full_name_no_prefix = row[5] if row[5] and isinstance(row[5], str) else val3

            # strip title prefix (นาย/นาง/นางสาว/เด็กชาย/เด็กหญิง)
            name_clean = re.sub(r'^(นาย|นางสาว|นาง|เด็กชาย|เด็กหญิง)\s*', '', full_name_no_prefix).strip()
            parts = name_clean.split()
            first_name = parts[0] if parts else name_clean
            last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

            if student_id in seen_ids or Student.objects.filter(student_id=student_id).exists():
                skipped += 1
                continue

            seen_ids.add(student_id)
            student = Student(
                student_id=student_id,
                first_name=first_name,
                last_name=last_name,
                phone='',
                grade=current_room,
            )
            student.set_password(student_id)
            student.save()
            created += 1

        self.stdout.write(self.style.SUCCESS(
            f'เพิ่มนักเรียนสำเร็จ {created} คน  (ข้ามที่มีอยู่แล้ว {skipped} คน)'
        ))
